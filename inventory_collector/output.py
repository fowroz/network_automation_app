"""
Report output -- CSV / XLSX / JSON, all with runtime-generated headers
(FR-3) and schema-drift handling for CSV/XLSX append mode (FR-3's
on_schema_change policy: new_file | expand | error).

Headers are NEVER hardcoded here: they're built from whatever `columns`
list the caller passes in (which the collector derives from the profile's
enabled fields at run time), so two different profiles naturally produce
two different, correct header rows (AC-8) with zero code changes.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

LOG = logging.getLogger("collector.output")

META_COLUMNS = ["TIMESTAMP", "TARGET_IP", "STATUS", "MISSING_FIELDS", "ERROR"]


class SchemaChangeError(RuntimeError):
    """Raised when on_schema_change: error is configured and the target
    file's existing header doesn't match the current run's columns."""


def build_headers(columns: list, include_metadata: bool = True) -> list:
    """The full ordered header row for a report: metadata columns first
    (if enabled), then every data column, in profile-declared order."""
    return ([*META_COLUMNS] if include_metadata else []) + list(columns)


def _timestamped_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def _read_existing_csv_header(path: Path) -> Optional[list]:
    if not path.exists() or path.stat().st_size == 0:
        return None
    with path.open(newline="", encoding="utf-8") as fh:
        try:
            return next(csv.reader(fh))
        except StopIteration:
            return None


def resolve_output_path_for_schema(
    path: Path,
    headers: list,
    mode: str = "append",
    on_schema_change: str = "new_file",
) -> tuple:
    """
    Implements FR-3's schema-drift policy for CSV/XLSX. Returns
    (final_path, final_headers, should_write_header_row).

    - mode == "overwrite": always (re)write from scratch with the given headers.
    - mode == "append", file doesn't exist yet: write fresh with given headers.
    - mode == "append", file exists with a MATCHING header: append, no new header row.
    - mode == "append", file exists with a DIFFERENT header:
        - on_schema_change == "new_file": redirect to a timestamped sibling file.
        - on_schema_change == "expand": keep old file, header becomes
          old-headers + any new columns not already present (existing rows'
          missing new columns are left blank by the CSV writer).
        - on_schema_change == "error": raise SchemaChangeError.
    """
    path = Path(path)
    if mode == "overwrite":
        return path, headers, True

    existing_header = _read_existing_csv_header(path)
    if existing_header is None:
        return path, headers, True

    if existing_header == headers:
        return path, headers, False

    LOG.warning("Schema drift detected for %s (existing header has %d column(s), "
                "current profile has %d) -- applying on_schema_change=%s",
                path, len(existing_header), len(headers), on_schema_change)

    if on_schema_change == "error":
        raise SchemaChangeError(
            f"CSV schema mismatch for {path}: existing header {existing_header} "
            f"does not match current profile's columns {headers}."
        )
    if on_schema_change == "expand":
        merged = list(existing_header) + [h for h in headers if h not in existing_header]
        return path, merged, "rewrite_needed"  # caller must rewrite existing rows against `merged`

    # default / "new_file"
    return _timestamped_path(path), headers, True


def write_csv(
    rows: list,
    columns: list,
    out_path: Path,
    mode: str = "append",
    include_metadata: bool = True,
    on_schema_change: str = "new_file",
) -> Path:
    """FR-3 (dynamic headers) + FR-3's schema-drift handling, AC-10
    (append preserves history: running twice produces 2x rows, header
    written once)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    headers = build_headers(columns, include_metadata)

    final_path, final_headers, header_action = resolve_output_path_for_schema(
        out_path, headers, mode=mode, on_schema_change=on_schema_change
    )

    if header_action == "rewrite_needed":
        # "expand" policy: re-read old rows, rewrite the whole file once
        # with the merged header so old rows keep their data and simply
        # gain blank cells for brand-new columns.
        old_rows = []
        if final_path.exists() and final_path.stat().st_size > 0:
            with final_path.open(newline="", encoding="utf-8") as fh:
                old_rows = list(csv.DictReader(fh))
        with final_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=final_headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(old_rows)
            writer.writerows(rows)
        LOG.info("Expanded schema and wrote %d new row(s) -> %s", len(rows), final_path)
        return final_path

    write_mode = "a" if (final_path.exists() and final_path.stat().st_size > 0 and header_action is False) else "w"
    with final_path.open(write_mode, newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=final_headers, extrasaction="ignore")
        if write_mode == "w":
            writer.writeheader()
        writer.writerows(rows)

    LOG.info("Wrote %d row(s) -> %s", len(rows), final_path)
    return final_path


def write_json(rows: list, out_path: Path, mode: str = "append") -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if mode == "append" and out_path.exists() and out_path.stat().st_size > 0:
        try:
            existing = json.loads(out_path.read_text(encoding="utf-8"))
            if not isinstance(existing, list):
                existing = [existing]
        except (json.JSONDecodeError, OSError):
            existing = []
        combined = existing + rows
    else:
        combined = rows

    out_path.write_text(json.dumps(combined, indent=2, default=str), encoding="utf-8")
    LOG.info("Wrote %d row(s) (%d total) -> %s", len(rows), len(combined), out_path)
    return out_path


def write_xlsx(
    rows: list,
    columns: list,
    out_path: Path,
    sheet_name: str = "Inventory",
    mode: str = "append",
    include_metadata: bool = True,
    on_schema_change: str = "new_file",
) -> Path:
    """XLSX output (one sheet per profile, per FR-7) using openpyxl.
    Applies the same schema-drift policy as CSV, keyed off the header row
    of the existing sheet (if any)."""
    from openpyxl import Workbook, load_workbook

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    headers = build_headers(columns, include_metadata)

    existing_header = None
    workbook = None
    if out_path.exists() and mode == "append":
        try:
            workbook = load_workbook(out_path)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                existing_header = list(first_row) if first_row else None
        except Exception as exc:
            LOG.warning("Could not read existing workbook %s (%s) -- starting fresh.", out_path, exc)
            workbook = None

    if mode == "overwrite" or workbook is None or existing_header is None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)
        final_path = out_path
        final_headers = headers
    elif existing_header == headers:
        sheet = workbook[sheet_name]
        final_path = out_path
        final_headers = headers
    else:
        LOG.warning("Schema drift detected for %s sheet '%s' -- applying on_schema_change=%s",
                    out_path, sheet_name, on_schema_change)
        if on_schema_change == "error":
            raise SchemaChangeError(
                f"XLSX schema mismatch for {out_path}::{sheet_name}: existing header "
                f"{existing_header} does not match current profile's columns {headers}."
            )
        if on_schema_change == "expand":
            final_headers = list(existing_header) + [h for h in headers if h not in existing_header]
            old_sheet = workbook[sheet_name]
            # Capture every existing data row (as dicts keyed by the OLD
            # header) before touching the sheet, so expanding the header
            # doesn't lose previously-collected data (mirrors write_csv's
            # "expand" behavior of rewriting old rows against the merged
            # header with blank cells for brand-new columns).
            old_rows = []
            for values in old_sheet.iter_rows(min_row=2, values_only=True):
                old_rows.append(dict(zip(existing_header, values)))
            workbook.remove(old_sheet)
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(final_headers)
            for old_row in old_rows:
                sheet.append([old_row.get(col, "") for col in final_headers])
            final_path = out_path
        else:  # new_file
            final_path = _timestamped_path(out_path)
            workbook = Workbook()
            workbook.remove(workbook.active)
            sheet = workbook.create_sheet(title=sheet_name)
            sheet.append(headers)
            final_headers = headers

    for row in rows:
        sheet.append([row.get(col, "") for col in final_headers])

    workbook.save(final_path)
    LOG.info("Wrote %d row(s) -> %s::%s", len(rows), final_path, sheet_name)
    return final_path


# ==========================================================================
# On-demand (in-memory) report generation -- no disk writes
# ==========================================================================
# The three functions above (write_csv/write_json/write_xlsx) are the
# ORIGINAL, still-supported disk-writing API used by the standalone
# collector.py CLI, which genuinely needs a persistent file on disk as
# its actual deliverable, plus append-mode history + schema-drift
# handling across many independent runs of the CLI over time.
#
# network_automation_app's web UI has a fundamentally different need:
# `audit_runs.report_json` in automation_console.db is ALREADY the
# durable source of truth for a completed run (see storage.save_audit_run()),
# so writing a SECOND, redundant copy to ./reports/*.csv on every single
# run -- scheduled or manual -- just accumulates disk usage forever with
# no corresponding benefit (nothing ever reads those files except the
# one-time "download this report" click, which can just as easily be
# generated fresh, on demand, from the JSON already sitting in SQLite).
# These render_*_bytes() functions produce the exact same byte-for-byte
# CSV/XLSX/JSON output as the disk-writing functions above (same
# build_headers() call, same row-to-cell mapping) but into an in-memory
# buffer (io.StringIO/io.BytesIO) instead of a file -- used by
# /audit/history/<id>/download (see app.py) to generate a downloadable
# file on the fly from a stored report_json, with no ./reports/ file
# ever created for that download in the first place.
def render_csv_bytes(rows: list, columns: list, include_metadata: bool = True) -> bytes:
    """Renders rows/columns to CSV format entirely in memory. Always a
    fresh, complete, single-shot render (no append/schema-drift concerns
    -- there is no pre-existing file to reconcile against)."""
    headers = build_headers(columns, include_metadata)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def render_json_bytes(rows: list) -> bytes:
    """Renders rows to JSON format entirely in memory."""
    return json.dumps(rows, indent=2, default=str).encode("utf-8")


def render_xlsx_bytes(rows: list, columns: list, sheet_name: str = "Report", include_metadata: bool = True) -> bytes:
    """Renders rows/columns to an XLSX workbook entirely in memory."""
    from openpyxl import Workbook

    headers = build_headers(columns, include_metadata)
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet(title=(sheet_name or "Report")[:31])  # Excel sheet-name length limit
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(col, "") for col in headers])

    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()
